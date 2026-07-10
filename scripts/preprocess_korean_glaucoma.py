#!/usr/bin/env python3
"""
파일명: scripts/preprocess_korean_glaucoma.py
목적:   한국인 녹내장 안저사진(수정본) 전처리 파이프라인

IRB: 국내 임상기관 IRB 승인 (2019)
보관: GPU 서버 로컬 전용 — 외부 반출 금지

입력: /dataset/korean_fundus_input/glaucoma_modified/
      (1.jpg~173.jpg + glaucoma_modified_info.xlsx)
출력: /dataset/korean_glaucoma_fundus/modified/

크롭 원칙: crop_layout_analysis.json에서 crop_possible=True인 파일만 처리.
           unknown 레이아웃은 skip_log.json에 기록 후 스킵.
"""
from __future__ import annotations

import argparse
import csv
import json
import sys
from datetime import datetime
from pathlib import Path

import cv2
import numpy as np
from openpyxl import load_workbook

_SCRIPTS = Path(__file__).resolve().parent
_ROOT = _SCRIPTS.parent
if str(_SCRIPTS) not in sys.path:
    sys.path.insert(0, str(_SCRIPTS))
from korean_gl_crop_utils import (
    analyze_bottom_layout,
    detect_boundaries,
    layout_to_color_boxes,
)

IRB_INFO = {
    "institution": "Korean Clinical Institution",
    "approved_year": 2019,
    "storage_policy": "LOCAL_ONLY",
    "external_transfer": "PROHIBITED",
    "git_commit": "PROHIBITED",
}

DATASET_PREFIX = "MEDI_KR_GL"
SUBSET_NAME = "modified"
OUTPUT_SIZE = (512, 512)
JPEG_QUALITY = 92
GREEN_G_MIN = 80
GREEN_RATIO = 1.4
MASK_PAD = 8

DEFAULT_INPUT = Path("/dataset/korean_fundus_input/glaucoma_modified")
DEFAULT_XLSX = DEFAULT_INPUT / "glaucoma_modified_info.xlsx"
DEFAULT_OUTPUT = Path("/dataset/korean_glaucoma_fundus")
DEFAULT_LAYOUT_JSON = _ROOT / "crop_layout_analysis.json"


def write_gitignore(output_dir: Path) -> None:
    gitignore_path = output_dir / ".gitignore"
    if not gitignore_path.exists():
        gitignore_path.write_text(
            "# MEDI-IoT Korean glaucoma clinical data\n"
            "# IRB 2019 — local only, no git commit\n"
            "*\n",
            encoding="utf-8",
        )


def load_label_xlsx(xlsx_path: Path) -> dict[int, dict]:
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active
    labels: dict[int, dict] = {}
    header = None
    for row in ws.iter_rows(values_only=True):
        if header is None:
            header = [str(c).strip() if c else "" for c in row]
            continue
        if not row[0]:
            continue
        img_no = int(row[1]) if row[1] else None
        rl = str(row[2]).strip().upper() if row[2] else None
        if img_no is None or rl not in ("R", "L"):
            continue
        entry = {
            "unit_no": str(row[3]).strip() if row[3] else "",
            "grade": int(row[4]) if row[4] else 0,
            "diagnosis": str(row[5]).strip() if row[5] else "",
            "diag_code": str(row[6]).strip() if row[6] else "",
            "age_raw": str(row[7]).strip() if row[7] else "",
            "sex": str(row[8]).strip().upper() if row[8] else "",
        }
        labels.setdefault(img_no, {})[rl] = entry
    wb.close()
    return labels


def load_layout_index(layout_path: Path) -> dict[str, dict]:
    if not layout_path.is_file():
        return {}
    data = json.loads(layout_path.read_text(encoding="utf-8"))
    index: dict[str, dict] = {}
    for entry in data.get("files") or []:
        key = str(entry.get("key") or entry.get("img_no", ""))
        index[key] = entry
    return index


def mask_pii(img: np.ndarray) -> tuple[np.ndarray, int]:
    result = img.copy()
    r = img[:, :, 2].astype(float)
    g = img[:, :, 1].astype(float)
    b = img[:, :, 0].astype(float)
    green = (g > GREEN_G_MIN) & (g > r * GREEN_RATIO) & (g > b * GREEN_RATIO)
    original_count = int(green.sum())
    kernel = np.ones((MASK_PAD * 2 + 1, MASK_PAD * 2 + 1), np.uint8)
    expanded = cv2.dilate(green.astype(np.uint8), kernel).astype(bool)
    result[expanded] = 0
    return result, original_count


def apply_clahe(img_bgr: np.ndarray) -> np.ndarray:
    lab = cv2.cvtColor(img_bgr, cv2.COLOR_BGR2LAB)
    l, a, b = cv2.split(lab)
    clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8, 8))
    return cv2.cvtColor(cv2.merge([clahe.apply(l), a, b]), cv2.COLOR_LAB2BGR)


def save_crop(
    img_masked: np.ndarray,
    box: tuple[int, int, int, int],
    out_path: Path,
    do_clahe: bool = True,
) -> None:
    y1, x1, y2, x2 = box
    crop = img_masked[y1:y2, x1:x2].copy()
    processed = apply_clahe(crop) if do_clahe else crop
    resized = cv2.resize(processed, OUTPUT_SIZE, interpolation=cv2.INTER_LANCZOS4)
    cv2.imwrite(str(out_path), resized, [cv2.IMWRITE_JPEG_QUALITY, JPEG_QUALITY])


def process_dataset(
    input_dir: Path,
    xlsx_path: Path,
    output_dir: Path,
    *,
    layout_path: Path,
    dry_run: bool = False,
    save_ir: bool = True,
) -> None:
    subset_root = output_dir / SUBSET_NAME
    print("=" * 60)
    print("MEDI-IoT Korean glaucoma modified preprocessing")
    print(f"IRB: {IRB_INFO['institution']} ({IRB_INFO['approved_year']})")
    if dry_run:
        print("*** DRY-RUN — no file writes ***")
    print(f"Input:  {input_dir}")
    print(f"Layout: {layout_path}")
    print(f"Output: {subset_root}/")
    print()

    dirs = {
        "color_od": subset_root / "color" / "OD",
        "color_os": subset_root / "color" / "OS",
        "ir_od": subset_root / "ir" / "OD",
        "ir_os": subset_root / "ir" / "OS",
        "logs": output_dir / "logs",
    }
    if not dry_run:
        for d in dirs.values():
            d.mkdir(parents=True, exist_ok=True)
        write_gitignore(output_dir)

    labels = load_label_xlsx(xlsx_path)
    layout_index = load_layout_index(layout_path)
    if layout_index:
        print(f"Layout index: {len(layout_index)} entries from {layout_path.name}")
    else:
        print(f"WARN: no layout JSON — inline analyze per image ({layout_path})")

    jpg_files = sorted(
        [f for f in input_dir.glob("*.jpg") if f.stem.isdigit()],
        key=lambda f: int(f.stem),
    )

    records: list[dict] = []
    skip_log: list[dict] = []
    stats = {
        "total": 0,
        "success": 0,
        "skip_no_label": 0,
        "skip_layout": 0,
        "error": 0,
        "layout_2split": 0,
        "layout_4split": 0,
        "grade": {},
        "diagnosis": {},
        "masking": {"total_green_pixels": 0, "files_with_pii": 0},
    }

    for jpg_file in jpg_files:
        img_no = int(jpg_file.stem)
        key = str(img_no)
        stats["total"] += 1
        if img_no not in labels:
            stats["skip_no_label"] += 1
            continue

        img_bgr = cv2.imread(str(jpg_file))
        if img_bgr is None:
            stats["error"] += 1
            continue

        layout_entry = layout_index.get(key)
        if layout_entry is None:
            layout_entry = analyze_bottom_layout(img_bgr)

        if not layout_entry.get("crop_possible"):
            stats["skip_layout"] += 1
            skip_log.append({
                "image_no": img_no,
                "path": str(jpg_file),
                "layout": layout_entry.get("layout"),
                "bottom_splits": layout_entry.get("bottom_splits"),
                "reason": layout_entry.get("reason", "layout_unknown"),
            })
            print(
                f"  [{img_no}] SKIP layout={layout_entry.get('layout')} "
                f"splits={layout_entry.get('bottom_splits')}"
            )
            continue

        color_boxes = layout_to_color_boxes(layout_entry)
        if not color_boxes:
            stats["skip_layout"] += 1
            continue

        _, w = img_bgr.shape[:2]
        split_row = int(layout_entry["split_row"])
        split_col = int(layout_entry["split_col"])
        layout_type = layout_entry.get("layout", "unknown")
        if layout_type == "2split":
            stats["layout_2split"] += 1
        elif layout_type == "4split":
            stats["layout_4split"] += 1

        print(
            f"  [{img_no}] {layout_type} 상하={split_row} splits={layout_entry.get('bottom_splits')} "
            f"OD={layout_entry.get('od_box')} OS={layout_entry.get('os_box')}"
        )

        img_masked, pii_count = mask_pii(img_bgr)
        if pii_count > 0:
            stats["masking"]["total_green_pixels"] += pii_count
            stats["masking"]["files_with_pii"] += 1

        crops: dict[str, dict] = {}
        if save_ir:
            crops["ir_R"] = {
                "box": (0, 0, split_row, split_col),
                "eye": "R",
                "modality": "ir",
                "dir": dirs["ir_od"],
            }
            crops["ir_L"] = {
                "box": (0, split_col, split_row, w),
                "eye": "L",
                "modality": "ir",
                "dir": dirs["ir_os"],
            }
        crops["color_R"] = {
            "box": color_boxes["color_R"],
            "eye": "R",
            "modality": "color",
            "dir": dirs["color_od"],
        }
        crops["color_L"] = {
            "box": color_boxes["color_L"],
            "eye": "L",
            "modality": "color",
            "dir": dirs["color_os"],
        }

        for cfg in crops.values():
            eye_key = cfg["eye"]
            if eye_key not in labels[img_no]:
                continue
            lbl = labels[img_no][eye_key]
            fname = (
                f"{DATASET_PREFIX}_{SUBSET_NAME}_{img_no:04d}_"
                f"{eye_key}_{cfg['modality']}.jpg"
            )
            out_path = cfg["dir"] / fname
            if not dry_run:
                try:
                    save_crop(
                        img_masked,
                        box=cfg["box"],
                        out_path=out_path,
                        do_clahe=(cfg["modality"] == "color"),
                    )
                except Exception as exc:
                    stats["error"] += 1
                    print(f"  error {fname}: {exc}")
                    continue

            g = lbl["grade"]
            d = lbl["diagnosis"]
            records.append({
                "image_no": img_no,
                "filename": fname,
                "subset": SUBSET_NAME,
                "eye": eye_key,
                "modality": cfg["modality"],
                "layout": layout_type,
                "_unit_no": lbl["unit_no"],
                "_age_raw": lbl["age_raw"],
                "_sex": lbl["sex"],
                "grade": g,
                "diagnosis": d,
                "diag_code": lbl["diag_code"],
                "glaucoma": 1,
                "glaucoma_grade": g,
                "is_ntg": 1 if "NTG" in d.upper() else 0,
                "is_poag": 1 if "POAG" in d.upper() else 0,
                "is_pacg": 1 if any(x in d.upper() for x in ("PACG", "CNAG")) else 0,
                "pii_masked": 1,
                "processed_at": datetime.now().isoformat(),
            })
            stats["success"] += 1
            stats["grade"][g] = stats["grade"].get(g, 0) + 1
            stats["diagnosis"][d] = stats["diagnosis"].get(d, 0) + 1

    if not dry_run:
        skip_path = output_dir / "skip_log.json"
        skip_path.write_text(
            json.dumps(
                {
                    "subset": SUBSET_NAME,
                    "skipped": len(skip_log),
                    "skip_ratio_pct": round(
                        stats["skip_layout"] / max(stats["total"], 1) * 100, 1
                    ),
                    "entries": skip_log,
                },
                indent=2,
                ensure_ascii=False,
            ),
            encoding="utf-8",
        )
        print(f"\nSkip log: {skip_path} ({len(skip_log)} entries)")

    if not dry_run and records:
        csv_path = output_dir / "labels_modified.csv"
        with open(csv_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=records[0].keys())
            writer.writeheader()
            writer.writerows(records)
        manifest = {
            "dataset": DATASET_PREFIX,
            "subset": SUBSET_NAME,
            "irb": IRB_INFO,
            "processed_at": datetime.now().isoformat(),
            "layout_json": str(layout_path),
            "total": len(records),
            "stats": stats,
            "source": "Korean Clinical Institution",
        }
        manifest_path = output_dir / "manifest_modified.json"
        with open(manifest_path, "w", encoding="utf-8") as f:
            json.dump(manifest, f, ensure_ascii=False, indent=2)
        print(f"Saved: {csv_path} ({len(records)} rows)")

    print()
    print(
        f"Total: {stats['total']}  OK: {stats['success']}  "
        f"skip_layout: {stats['skip_layout']}  no_label: {stats['skip_no_label']}  "
        f"err: {stats['error']}"
    )
    if stats["total"]:
        print(
            f"Layout: 2split={stats['layout_2split']} 4split={stats['layout_4split']} "
            f"skip={stats['skip_layout']} ({stats['skip_layout']/stats['total']*100:.1f}%)"
        )


def main() -> None:
    p = argparse.ArgumentParser()
    p.add_argument("--input-dir", type=Path, default=DEFAULT_INPUT)
    p.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
    p.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT)
    p.add_argument("--layout-json", type=Path, default=DEFAULT_LAYOUT_JSON)
    p.add_argument("--dry-run", action="store_true")
    p.add_argument("--no-ir", action="store_true")
    args = p.parse_args()

    if not args.input_dir.exists():
        print(f"error: missing {args.input_dir}", file=sys.stderr)
        sys.exit(1)
    if not args.xlsx.exists():
        print(f"error: missing {args.xlsx}", file=sys.stderr)
        sys.exit(1)

    layout_path = args.layout_json
    if not layout_path.is_absolute():
        layout_path = _ROOT / layout_path

    process_dataset(
        input_dir=args.input_dir,
        xlsx_path=args.xlsx,
        output_dir=args.output_dir,
        layout_path=layout_path,
        dry_run=args.dry_run,
        save_ir=not args.no_ir,
    )


if __name__ == "__main__":
    main()
