#!/usr/bin/env python3
"""
파일명: scripts/preprocess_korean_gl_origin.py
목적:   한국인 녹내장 원본 데이터(glaucoma_origin) 전처리 파이프라인

IRB: 국내 임상기관 IRB 승인 (2019)
보관: GPU 서버 로컬 전용 — 외부 반출 금지

입력 구조 (glaucoma_origin/):
  {folder_no}/
    {n}_YYYYMMDD_Optic_Nerve_Fiber_Layer_Photography+
                 Optic_Disc_Stereophotography_양측_{i}.jpg  ← 안저사진 합본
    {n}_YYYYMMDD_Sita_Visual_Field_ExaminationAutomated_양측_8.jpg  ← 시야검사 OD
    {n}_YYYYMMDD_Sita_Visual_Field_ExaminationAutomated_양측_9.jpg  ← 시야검사 OS
    {n}_YYYYMMDD_OCT_녹내장(시신경)_양안_{i}.jpg                    ← OCT 결과지

출력 구조 (korean_glaucoma_fundus/origin/):
  fundus/OD/  MEDI_KR_GL_orig_{folder:04d}_YYYYMMDD_R_color.jpg
  fundus/OS/  MEDI_KR_GL_orig_{folder:04d}_YYYYMMDD_L_color.jpg
  fundus/ir/OD/  (적외선)
  fundus/ir/OS/
  vf/OD/      MEDI_KR_GL_orig_{folder:04d}_YYYYMMDD_R_vf.jpg   ← 시야검사(원본 보존)
  vf/OS/      MEDI_KR_GL_orig_{folder:04d}_YYYYMMDD_L_vf.jpg
  oct/        MEDI_KR_GL_orig_{folder:04d}_YYYYMMDD_oct.jpg    ← OCT(원본 보존)
  labels.csv
  manifest.json

히스토리:
  2026-07-07 - 최초 작성
"""
from __future__ import annotations

import argparse
import csv
import json
import re
import sys
from datetime import datetime
from pathlib import Path

import cv2
import numpy as np
from openpyxl import load_workbook

# ── 상수 ──────────────────────────────────────────────────────────────────────
DATASET_PREFIX = "MEDI_KR_GL"
OUTPUT_SIZE    = (512, 512)    # 안저사진 출력 크기
VF_SIZE        = (1024, 768)   # 시야검사 출력 크기 (수치 가독성 보존)
OCT_SIZE       = (1200, 900)   # OCT 출력 크기 (구조 보존)
JPEG_QUALITY   = 92

# 파일 종류 구분 키워드
KEYWORDS = {
    "fundus": ["optic_nerve", "optic nerve", "stereophotography",
               "opticnerve", "disc", "시신경유두", "안저"],
    "vf":     ["sita", "visual_field", "visual field", "시야"],
    "oct":    ["oct", "rnfl", "녹내장"],
}

DEFAULT_INPUT = Path("/dataset/korean_fundus_input/glaucoma_origin")
DEFAULT_XLSX = DEFAULT_INPUT / "glaucoma_origin_info.xlsx"
DEFAULT_OUTPUT = Path("/dataset/korean_glaucoma_fundus")

# 마스킹 파라미터
GREEN_G_MIN  = 80
GREEN_RATIO  = 1.4
MASK_PAD     = 8

IRB_INFO = {
    "institution": "Korean Clinical Institution",
    "approved_year": 2019,
    "storage_policy": "LOCAL_ONLY",
    "external_transfer": "PROHIBITED",
    "git_commit": "PROHIBITED",
}


# ── 유틸 ──────────────────────────────────────────────────────────────────────

def classify_file(fname: str) -> str:
    """
    파일명 키워드로 종류 분류
    반환: 'fundus' | 'vf' | 'oct' | 'unknown'
    """
    lower = fname.lower().replace(" ", "_")
    for kind, kws in KEYWORDS.items():
        for kw in kws:
            if kw.replace(" ", "_") in lower:
                return kind
    return "unknown"


def extract_date(fname: str) -> str:
    """파일명에서 날짜 추출 (YYYYMMDD)"""
    m = re.search(r"(\d{8})", fname)
    return m.group(1) if m else "00000000"


def classify_vf_eye(fname: str) -> str:
    """
    시야검사 파일의 안구 판별
    파일명 끝 숫자: _8 = OD(우안), _9 = OS(좌안) (관찰된 패턴)
    """
    lower = fname.lower()
    m = re.search(r"_(\d+)\.jpg$", lower)
    if m:
        n = int(m.group(1))
        if n % 2 == 0:   # 짝수 → OD
            return "R"
        else:             # 홀수 → OS
            return "L"
    return "R"  # 기본값


def mask_pii(img: np.ndarray) -> tuple[np.ndarray, int]:
    """초록색 개인정보 텍스트 마스킹"""
    result = img.copy()
    r = img[:, :, 2].astype(float)
    g = img[:, :, 1].astype(float)
    b = img[:, :, 0].astype(float)
    green = (g > GREEN_G_MIN) & (g > r * GREEN_RATIO) & (g > b * GREEN_RATIO)
    count = int(green.sum())
    kernel = np.ones((MASK_PAD * 2 + 1, MASK_PAD * 2 + 1), np.uint8)
    expanded = cv2.dilate(green.astype(np.uint8), kernel).astype(bool)
    result[expanded] = 0
    return result, count


def mask_text_regions(img: np.ndarray) -> np.ndarray:
    """
    시야검사/OCT 결과지의 텍스트 영역 마스킹
    - 초록색 텍스트 (안저사진 장비 정보)
    - 검은 배경의 흰색 텍스트는 임상 수치이므로 보존
    - 상단 헤더(환자명/ID 있는 부분)만 추가 마스킹
    """
    result, _ = mask_pii(img)  # 초록 텍스트 우선 제거

    # 시야검사/OCT는 상단 헤더에 이름/ID 있음 → 상단 15% 마스킹
    h = img.shape[0]
    header_height = int(h * 0.15)
    result[:header_height, :] = 0

    return result


def detect_fundus_boundaries(img: np.ndarray) -> tuple[int, int]:
    """안저사진 합본 상하/좌우 경계 자동 감지"""
    h, w = img.shape[:2]
    row_means = img.mean(axis=(1, 2))
    s, e = int(h * 0.30), int(h * 0.70)
    split_row = s + int(row_means[s:e].argmin())

    col_means = img[:split_row].mean(axis=(0, 2))
    cs, ce = int(w * 0.30), int(w * 0.70)
    split_col = cs + int(col_means[cs:ce].argmin())

    return split_row, split_col


def apply_clahe(img_bgr: np.ndarray) -> np.ndarray:
    """CLAHE 전처리"""
    lab = cv2.cvtColor(img_bgr, cv2.COLOR_BGR2LAB)
    l, a, b = cv2.split(lab)
    clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8, 8))
    return cv2.cvtColor(cv2.merge([clahe.apply(l), a, b]), cv2.COLOR_LAB2BGR)


def load_label_xlsx(xlsx_path: Path) -> dict[int, dict]:
    """정보.xlsx → {Folder_No: {R: {...}, L: {...}}}"""
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active
    labels: dict[int, dict] = {}
    header = None

    for row in ws.iter_rows(values_only=True):
        if header is None:
            header = [str(c).strip() if c else "" for c in row]
            continue
        if not row[0]:
            continue

        # Folder_No는 컬럼 1 (Image_No와 동일 위치)
        folder_no = int(row[1]) if row[1] else None
        rl = str(row[2]).strip().upper() if row[2] else None
        if folder_no is None or rl not in ("R", "L"):
            continue

        entry = {
            "unit_no":   str(row[3]).strip() if row[3] else "",
            "grade":     int(row[4]) if row[4] else 0,
            "diagnosis": str(row[5]).strip() if row[5] else "",
            "diag_code": str(row[6]).strip() if row[6] else "",
            "age_raw":   str(row[7]).strip() if row[7] else "",
            "sex":       str(row[8]).strip().upper() if row[8] else "",
        }
        if folder_no not in labels:
            labels[folder_no] = {}
        labels[folder_no][rl] = entry

    wb.close()
    return labels


def write_gitignore(output_dir: Path) -> None:
    gi = output_dir / ".gitignore"
    if not gi.exists():
        gi.write_text("*\n# IRB 2019 — 로컬 전용, Git 커밋 금지\n")


# ── 메인 파이프라인 ────────────────────────────────────────────────────────────

def process_origin(
    input_dir: Path,
    xlsx_path: Path,
    output_dir: Path,
    dry_run: bool = False,
) -> None:
    print("=" * 60)
    print("MEDI-IoT 한국인 녹내장 원본 데이터 전처리")
    print(f"IRB: {IRB_INFO['institution']} ({IRB_INFO['approved_year']})")
    print("=" * 60)

    # 출력 디렉토리
    dirs = {
        "fundus_od":  output_dir / "origin" / "fundus" / "OD",
        "fundus_os":  output_dir / "origin" / "fundus" / "OS",
        "fundus_ir_od": output_dir / "origin" / "fundus" / "ir" / "OD",
        "fundus_ir_os": output_dir / "origin" / "fundus" / "ir" / "OS",
        "vf_od":      output_dir / "origin" / "vf" / "OD",
        "vf_os":      output_dir / "origin" / "vf" / "OS",
        "oct":        output_dir / "origin" / "oct",
        "logs":       output_dir / "logs",
    }
    if not dry_run:
        for d in dirs.values():
            d.mkdir(parents=True, exist_ok=True)
        write_gitignore(output_dir)

    # 라벨 로드
    labels = load_label_xlsx(xlsx_path)
    print(f"라벨: {len(labels)}개 Folder_No")

    # 폴더 목록 (숫자 폴더만)
    folders = sorted(
        [f for f in input_dir.iterdir()
         if f.is_dir() and f.name.isdigit()],
        key=lambda f: int(f.name)
    )
    print(f"폴더: {len(folders)}개")
    print()

    records = []
    stats = {
        "total_folders": 0,
        "fundus_saved":  0,
        "vf_saved":      0,
        "oct_saved":     0,
        "multi_visit":   0,   # 복수 방문 있는 폴더
        "error":         0,
        "grade": {}, "diagnosis": {},
    }
    errors = []

    for folder in folders:
        folder_no = int(folder.name)
        stats["total_folders"] += 1

        lbl_both = labels.get(folder_no, {})
        if not lbl_both:
            print(f"  ⚠️  폴더 {folder_no}: 라벨 없음 → 스킵")
            continue

        # 폴더 내 파일 분류
        files = sorted(folder.glob("*.jpg"))
        by_kind: dict[str, list[Path]] = {
            "fundus": [], "vf": [], "oct": [], "unknown": []
        }
        dates_seen: set[str] = set()
        for f in files:
            kind = classify_file(f.name)
            by_kind[kind].append(f)
            dates_seen.add(extract_date(f.name))

        if len(dates_seen) > 1:
            stats["multi_visit"] += 1

        print(f"  [{folder_no:3d}] 안저={len(by_kind['fundus'])} "
              f"시야={len(by_kind['vf'])} OCT={len(by_kind['oct'])} "
              f"방문={len(dates_seen)}회")

        # ── 안저사진 처리 ────────────────────────────────────────────────────
        for fundus_file in by_kind["fundus"]:
            date_str = extract_date(fundus_file.name)
            img_bgr = cv2.imread(str(fundus_file))
            if img_bgr is None:
                errors.append({"file": str(fundus_file), "reason": "load_failed"})
                continue

            img_masked, pii_cnt = mask_pii(img_bgr)
            split_row, split_col = detect_fundus_boundaries(img_bgr)
            h, w = img_bgr.shape[:2]

            crops = {
                "R_color": (split_row, split_col//2, h, split_col,
                            dirs["fundus_od"], True),
                "L_color": (split_row, split_col, h, split_col+(w-split_col)//2,
                            dirs["fundus_os"], True),
                "R_ir":    (0, 0, split_row, split_col,
                            dirs["fundus_ir_od"], False),
                "L_ir":    (0, split_col, split_row, w,
                            dirs["fundus_ir_os"], False),
            }

            for crop_key, (y1,x1,y2,x2, out_dir, do_clahe) in crops.items():
                eye = crop_key[0]  # R or L
                mod = crop_key[2:]  # color or ir
                lbl = lbl_both.get(eye, {})
                if not lbl:
                    continue

                fname = (f"{DATASET_PREFIX}_orig_{folder_no:04d}_"
                         f"{date_str}_{eye}_{mod}.jpg")
                out_path = out_dir / fname

                if not dry_run:
                    try:
                        crop = img_masked[y1:y2, x1:x2]
                        processed = apply_clahe(crop) if do_clahe else crop
                        resized = cv2.resize(processed, OUTPUT_SIZE,
                                             interpolation=cv2.INTER_LANCZOS4)
                        cv2.imwrite(str(out_path), resized,
                                    [cv2.IMWRITE_JPEG_QUALITY, JPEG_QUALITY])
                    except Exception as e:
                        errors.append({"file": fname, "reason": str(e)})
                        continue

                g = lbl.get("grade", 0)
                d = lbl.get("diagnosis", "")
                record = {
                    "folder_no":    folder_no,
                    "date":         date_str,
                    "filename":     fname,
                    "file_type":    "fundus",
                    "modality":     mod,
                    "eye":          eye,
                    "_unit_no":     lbl.get("unit_no",""),
                    "_age_raw":     lbl.get("age_raw",""),
                    "_sex":         lbl.get("sex",""),
                    "grade":        g,
                    "diagnosis":    d,
                    "diag_code":    lbl.get("diag_code",""),
                    "glaucoma":     1,
                    "glaucoma_grade": g,
                    "is_ntg":       1 if "NTG" in d.upper() else 0,
                    "is_poag":      1 if "POAG" in d.upper() else 0,
                    "is_pacg":      1 if any(x in d.upper()
                                            for x in ["PACG","CNAG"]) else 0,
                    "pii_masked":   1,
                    "visit_count":  len(dates_seen),
                    "processed_at": datetime.now().isoformat(),
                }
                records.append(record)
                stats["fundus_saved"] += 1
                stats["grade"][g] = stats["grade"].get(g, 0) + 1
                stats["diagnosis"][d] = stats["diagnosis"].get(d, 0) + 1

        # ── 시야검사 처리 ────────────────────────────────────────────────────
        for vf_file in by_kind["vf"]:
            date_str = extract_date(vf_file.name)
            eye = classify_vf_eye(vf_file.name)
            lbl = lbl_both.get(eye, {})
            if not lbl:
                continue

            img_bgr = cv2.imread(str(vf_file))
            if img_bgr is None:
                continue

            img_masked = mask_text_regions(img_bgr)

            fname = (f"{DATASET_PREFIX}_orig_{folder_no:04d}_"
                     f"{date_str}_{eye}_vf.jpg")
            out_dir = dirs["vf_od"] if eye == "R" else dirs["vf_os"]
            out_path = out_dir / fname

            if not dry_run:
                try:
                    resized = cv2.resize(img_masked, VF_SIZE,
                                         interpolation=cv2.INTER_LANCZOS4)
                    cv2.imwrite(str(out_path), resized,
                                [cv2.IMWRITE_JPEG_QUALITY, JPEG_QUALITY])
                except Exception as e:
                    errors.append({"file": fname, "reason": str(e)})
                    continue

            g = lbl.get("grade", 0)
            d = lbl.get("diagnosis", "")
            records.append({
                "folder_no":    folder_no,
                "date":         date_str,
                "filename":     fname,
                "file_type":    "vf",
                "modality":     "vf",
                "eye":          eye,
                "_unit_no":     lbl.get("unit_no",""),
                "_age_raw":     lbl.get("age_raw",""),
                "_sex":         lbl.get("sex",""),
                "grade":        g,
                "diagnosis":    d,
                "diag_code":    lbl.get("diag_code",""),
                "glaucoma":     1,
                "glaucoma_grade": g,
                "is_ntg":       1 if "NTG" in d.upper() else 0,
                "is_poag":      1 if "POAG" in d.upper() else 0,
                "is_pacg":      1 if any(x in d.upper()
                                        for x in ["PACG","CNAG"]) else 0,
                "pii_masked":   1,
                "visit_count":  len(dates_seen),
                "processed_at": datetime.now().isoformat(),
            })
            stats["vf_saved"] += 1

        # ── OCT 처리 ────────────────────────────────────────────────────────
        for oct_file in by_kind["oct"]:
            date_str = extract_date(oct_file.name)
            img_bgr = cv2.imread(str(oct_file))
            if img_bgr is None:
                continue

            img_masked = mask_text_regions(img_bgr)

            fname = (f"{DATASET_PREFIX}_orig_{folder_no:04d}_"
                     f"{date_str}_oct.jpg")
            out_path = dirs["oct"] / fname

            if not dry_run:
                try:
                    resized = cv2.resize(img_masked, OCT_SIZE,
                                         interpolation=cv2.INTER_LANCZOS4)
                    cv2.imwrite(str(out_path), resized,
                                [cv2.IMWRITE_JPEG_QUALITY, JPEG_QUALITY])
                except Exception as e:
                    errors.append({"file": fname, "reason": str(e)})
                    continue

            # OCT는 양안 동시 저장 — R 라벨 기준으로 1건만 기록
            lbl = lbl_both.get("R", lbl_both.get("L", {}))
            g = lbl.get("grade", 0)
            d = lbl.get("diagnosis", "")
            records.append({
                "folder_no":    folder_no,
                "date":         date_str,
                "filename":     fname,
                "file_type":    "oct",
                "modality":     "oct",
                "eye":          "OU",   # 양안
                "_unit_no":     lbl.get("unit_no",""),
                "_age_raw":     lbl.get("age_raw",""),
                "_sex":         lbl.get("sex",""),
                "grade":        g,
                "diagnosis":    d,
                "diag_code":    lbl.get("diag_code",""),
                "glaucoma":     1,
                "glaucoma_grade": g,
                "is_ntg":       1 if "NTG" in d.upper() else 0,
                "is_poag":      1 if "POAG" in d.upper() else 0,
                "is_pacg":      1 if any(x in d.upper()
                                        for x in ["PACG","CNAG"]) else 0,
                "pii_masked":   1,
                "visit_count":  len(dates_seen),
                "processed_at": datetime.now().isoformat(),
            })
            stats["oct_saved"] += 1

    # ── 저장 ─────────────────────────────────────────────────────────────────
    if not dry_run and records:
        csv_path = output_dir / "labels_origin.csv"
        with open(csv_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=records[0].keys())
            writer.writeheader()
            writer.writerows(records)

        manifest = {
            "dataset":      f"{DATASET_PREFIX}_ORIGIN",
            "irb":          IRB_INFO,
            "processed_at": datetime.now().isoformat(),
            "total_records": len(records),
            "stats":        stats,
            "file_types": {
                "fundus": "안저사진 (컬러/IR) — AI 훈련 직접 사용",
                "vf":     "시야검사 결과지 — 라벨 보강, 향후 OCR/분석 활용",
                "oct":    "OCT 결과지 — CDR/RNFL 수치 추출 후 라벨 보강 활용",
            },
            "WARNING": (
                "IRB 2019 승인. GPU 서버 로컬 전용. "
                "외부 반출/Git 커밋 금지."
            ),
        }
        manifest_path = output_dir / "manifest_origin.json"
        with open(manifest_path, "w", encoding="utf-8") as f:
            json.dump(manifest, f, ensure_ascii=False, indent=2)

        if errors:
            err_path = dirs["logs"] / "errors_origin.json"
            with open(err_path, "w", encoding="utf-8") as f:
                json.dump(errors, f, ensure_ascii=False, indent=2)

        print(f"\n저장 완료:")
        print(f"  CSV:      {csv_path} ({len(records)}행)")
        print(f"  manifest: {manifest_path}")

    # ── 통계 ─────────────────────────────────────────────────────────────────
    print()
    print("=" * 40)
    print("처리 결과")
    print("=" * 40)
    print(f"  폴더:       {stats['total_folders']}개")
    print(f"  안저사진:   {stats['fundus_saved']}장")
    print(f"  시야검사:   {stats['vf_saved']}장")
    print(f"  OCT:        {stats['oct_saved']}장")
    print(f"  복수방문:   {stats['multi_visit']}개 폴더")
    print(f"  에러:       {stats['error']}건")
    print()
    print("  Grade:")
    for g in sorted(stats["grade"]):
        s = {1:"경증",2:"중등도",3:"중증"}.get(g,"?")
        print(f"    Grade {g}({s}): {stats['grade'][g]}건")
    print()
    print("  진단:")
    for d, cnt in sorted(stats["diagnosis"].items(), key=lambda x:-x[1]):
        print(f"    {d}: {cnt}건")


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--input-dir", type=Path, default=DEFAULT_INPUT)
    p.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
    p.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT)
    p.add_argument("--dry-run",    action="store_true")
    args = p.parse_args()

    if not args.input_dir.exists():
        print(f"오류: {args.input_dir} 없음"); sys.exit(1)
    if not args.xlsx.exists():
        print(f"오류: {args.xlsx} 없음"); sys.exit(1)

    process_origin(
        input_dir=args.input_dir,
        xlsx_path=args.xlsx,
        output_dir=args.output_dir,
        dry_run=args.dry_run,
    )

if __name__ == "__main__":
    main()
